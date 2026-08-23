import csv
import io
import json
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.orm import Session
from app.db.session import get_db
from app.db.models import Product
from app.pipeline.unilog_formatter import UnilogFormatter, UNILOG_DELIVERY_HEADERS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter(prefix="/export", tags=["Export"])


@router.get("/catalog")
def export_full_catalog_delivery_format(
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    db: Session = Depends(get_db)
):
    """
    Exports all products in the database into the official 252-column UniHack Expected Output Delivery Format (CSV or XLSX).
    """
    products = db.query(Product).all()

    if format == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Delivery Format"

        # Styling definitions
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # Write 252 static headers
        ws.append(UNILOG_DELIVERY_HEADERS)
        for col_num in range(1, len(UNILOG_DELIVERY_HEADERS) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Write rows
        for p in products:
            row_dict = UnilogFormatter.format_product_row(p)
            row_values = [row_dict.get(h, "") for h in UNILOG_DELIVERY_HEADERS]
            ws.append(row_values)

        output_stream = io.BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)

        return StreamingResponse(
            output_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=Unihack_Enriched_Product_Delivery_Format.xlsx"
            }
        )

    # Default CSV streaming
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=UNILOG_DELIVERY_HEADERS)
    writer.writeheader()

    for p in products:
        row = UnilogFormatter.format_product_row(p)
        writer.writerow(row)

    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=Unihack_Enriched_Product_Delivery_Format.csv"
        }
    )


@router.get("/catalog/xlsx")
def export_catalog_xlsx(db: Session = Depends(get_db)):
    """Direct shortcut for downloading official 252-column XLSX."""
    return export_full_catalog_delivery_format(format="xlsx", db=db)


@router.get("/catalog/csv")
def export_catalog_csv(db: Session = Depends(get_db)):
    """Direct shortcut for downloading official 252-column CSV."""
    return export_full_catalog_delivery_format(format="csv", db=db)


@router.get("/products/{product_id}")
def export_product_data(
    product_id: str,
    format: str = Query("json", pattern="^(json|csv|delivery_format_csv|xlsx)$"),
    db: Session = Depends(get_db)
):
    """
    Exports product intelligence record in PIM-compatible JSON, standard CSV, or official 252-column UniHack delivery format.
    """
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Product with ID '{product_id}' not found."
        )

    if format == "delivery_format_csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=UNILOG_DELIVERY_HEADERS)
        writer.writeheader()
        row = UnilogFormatter.format_product_row(product)
        writer.writerow(row)

        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=product_{product.id}_unilog_delivery_format.csv"
            }
        )

    # Clean product payload
    export_payload = {
        "product_id": product.id,
        "product_name": product.name,
        "category": product.category,
        "manufacturer": product.manufacturer,
        "sku": product.sku,
        "health_score": product.health_score,
        "attributes": [
            {
                "name": a.name,
                "value": a.normalized_value or a.raw_value,
                "raw_value": a.raw_value,
                "unit": a.unit,
                "knowledge_type": a.knowledge_type.value,
                "trust_status": a.trust_status.value,
                "confidence": a.confidence,
                "is_inferred": a.is_inferred,
                "evidence_quote": a.evidence.text_quote if a.evidence else None,
                "page_number": a.evidence.page_number if a.evidence else None
            }
            for a in product.attributes
        ]
    }

    if format == "json":
        return JSONResponse(
            content=export_payload,
            headers={
                "Content-Disposition": f"attachment; filename=product_{product.id}_unilogger.json"
            }
        )

    # Standard CSV Export format
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "Product ID", "Product Name", "Category", "Manufacturer", "SKU",
        "Health Score", "Attribute Name", "Value", "Raw Value", "Unit",
        "Knowledge Type", "Trust Status", "Confidence", "Is Inferred",
        "Evidence Quote", "Page Number"
    ])

    for a in product.attributes:
        writer.writerow([
            product.id,
            product.name,
            product.category or "",
            product.manufacturer or "",
            product.sku or "",
            product.health_score,
            a.name,
            a.normalized_value or a.raw_value or "",
            a.raw_value or "",
            a.unit or "",
            a.knowledge_type.value,
            a.trust_status.value,
            a.confidence,
            a.is_inferred,
            a.evidence.text_quote if a.evidence else "",
            a.evidence.page_number if a.evidence else ""
        ])

    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=product_{product.id}_unilogger.csv"
        }
    )
